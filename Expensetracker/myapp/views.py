import pandas as pd
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
import numpy as np

def index(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            return HttpResponse("No file selected for uploading")
        
        file_size = excel_file.size
       # print(f"Uploaded file size: {file_size} bytes")

        try:
            # Load the workbook using openpyxl
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active

            # Initialize variables for keyword search
            keyword_locations_pf_monthly = []
            keyword_locations_pf_budget = []

            # Search for 'P.F. Budget' and 'P.F. Budget Utlised' in the worksheet
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and "P.F. Budget".lower() in str(cell.value).lower():
                        keyword_locations_pf_monthly.append((cell.row, cell.column, cell.value))
                       # print(f"Found 'P.F. Budget' at row {cell.row}, column {openpyxl.utils.get_column_letter(cell.column)}")
                    if cell.value and "P.F. Budget Utlised".lower() in str(cell.value).lower():
                        keyword_locations_pf_budget.append((cell.row, cell.column, cell.value))
                      #  print(f"Found 'P.F. Budget Utlised' at row {cell.row}, column {openpyxl.utils.get_column_letter(cell.column)}")

            # Output the list of found locations for 'P.F. Budget'
           # print("Locations found for 'P.F. Budget':")
            for location in keyword_locations_pf_monthly:
                row_num, col_num, cell_value = location
             #   print(f"Row: {row_num}, Value: {cell_value}")

            # Output the list of found locations for 'P.F. Budget Utlised'
            #print("Locations found for 'P.F. Budget Utlised':")
            for location in keyword_locations_pf_budget:
                row_num, col_num, cell_value = location
             #   print(f"Row: {row_num}, Value: {cell_value}")

            # Read the uploaded Excel file into a Pandas DataFrame
            df = pd.read_excel(excel_file, sheet_name=None, index_col=-1)  # Load all sheets as a dictionary of DataFrames
            sheet1_df = df['Sheet1']
            # Optionally, you can clean up NaN values if needed
            sheet1_df = sheet1_df.dropna(how='all')  # Drop rows where all values are NaN
            # Print the DataFrame for debugging
            #print("DataFrame for Sheet1:")
           # print(sheet1_df)
            
            # Process each sheet in the workbook
            alert_messages = []
            excel_data = []

            for sheet_name, data in df.items():
                data = data.fillna('')  # Replace NaN with empty strings for display purposes
                # Convert the DataFrame to a list of lists (each list representing a row)
                sheet_data = data.values.tolist()

                # Add headers separately
                headers = data.columns.tolist()
                excel_data.append({"sheet_name": sheet_name, "headers": headers, "data": sheet_data})

                # Initialize lists to store values
                pf_Budget_values = []
                pf_budget_utlised = []
                travel_Budget_values = []
                travel_budget_utlised = []
                consultant_code_values = []
                travel_days_Budget_values = []
                travel_days_budget_utlised = []
                dsa_budget_values = []
                dsa_budget_utlised = []
                dsa_days_budget_values = []
                dsa_days_budget_utlised = []
                terminal_charges_Budget_values = []
                terminal_charges_budget_utlised = []




                # Check the 'Consultant Code' column
                if "Consultant Code" in headers:
                    consultant_code_index = headers.index("Consultant Code")
                    consultant_code_values = [str(row[consultant_code_index]).replace('\xa0', '').strip() for row in sheet_data]
                    # Print Consultant Code values
                   # print(f"Consultant Code values in sheet '{sheet_name}': {consultant_code_values}")

                # Check the 'P.F. Budget' column
                if "P.F. Budget" in headers:
                    pf_index = headers.index("P.F. Budget")
                    pf_Budget_values = [str(row[pf_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                    # Print P.F. Budget values
                   # print(f"P.F. Budget values in sheet '{sheet_name}': {pf_Budget_values}")

                # Check the 'P.F. Budget Utlised' column
                if "P.F. Budget Utlised" in headers:
                    pf_index = headers.index("P.F. Budget Utlised")
                    pf_budget_utlised = [str(row[pf_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                    # Print P.F. Budget Utlised values
                   # print(f"P.F. Budget Utlised values in sheet '{sheet_name}': {pf_budget_utlised}")
                    
                    # Subtract P.F. Budget Utlised from P.F. Budget
                    pf_subtracted_values = []
                    for pf_monthly, pf_budget_utlised in zip(pf_Budget_values, pf_budget_utlised):
                        if pf_monthly.strip() and pf_budget_utlised.strip():  # Check if both values are not empty
                            try:
                                # Remove commas before converting to float
                                pf_monthly_float = float(pf_monthly.replace(',', ''))
                                pf_budget_utlised_float = float(pf_budget_utlised.replace(',', ''))
                                pf_subtracted_values.append(pf_monthly_float - pf_budget_utlised_float)
                            except ValueError as e:
                              #  print(f"Error converting values: {e}")
                                pf_subtracted_values.append('')  # or some other value to indicate an error
                        else:
                            pf_subtracted_values.append('')  # or some other value to indicate an error
                    # Print the subtracted values
                    # print(f"Subtracted values (P.F. Budget): {pf_subtracted_values}")
                    
                # Check the 'Travel Budget' column
                if "Travel Budget" in headers:
                    travel_index = headers.index("Travel Budget")
                    travel_Budget_values = [str(row[travel_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                    # Print Travel Budget values
                    # print(f"Travel Budget values in sheet '{sheet_name}': {travel_Budget_values}")

                # Check the 'Travel Budget Utlised' column
                if "Travel Budget Utlised" in headers:
                    travel_index = headers.index("Travel Budget Utlised")
                    travel_budget_utlised = [str(row[travel_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                    # Print Travel Budget Utlised values
                    # print(f"Travel Budget Utlised values in sheet '{sheet_name}': {travel_budget_utlised}")
                
                    # Subtract Travel Budget Utlised from Travel Budget
                    travel_subtracted_values = []
                    for travel_monthly, travel_budget_utlised in zip(travel_Budget_values, travel_budget_utlised):
                        if travel_monthly.strip() and travel_budget_utlised.strip():  # Check if both values are not empty
                            try:
                                # Remove commas before converting to float
                                travel_monthly_float = float(travel_monthly.replace(',', ''))
                                travel_budget_utlised_float = float(travel_budget_utlised.replace(',', ''))
                                travel_subtracted_values.append(travel_monthly_float - travel_budget_utlised_float)
                            except ValueError as e:
                               # print(f"Error converting values: {e}")
                                travel_subtracted_values.append('')  # or some other value to indicate an error
                        else:
                            travel_subtracted_values.append('')  # or some other value to indicate an error
                    # Print the subtracted values
                    # print(f"Subtracted values (Travel Budget): {travel_subtracted_values}")


                # Check the 'Travel Days Budget' column
                if "Travel Days Budget" in headers:
                    travel_days_index = headers.index("Travel Days Budget")
                    travel_days_Budget_values = [str(row[travel_days_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values
                
                # Check the 'Travel Days Budget Utlised' column
                if "Travel Days Budget Utlised" in headers:
                    travel_days_index = headers.index("Travel Days Budget Utlised")
                    travel_days_budget_utlised = [str(row[travel_days_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values
                    
                    # Subtract Travel Days Budget Utlised from Travel Days Budget
                    travel_days_subtracted_values = []
                    for travel_days_budget, travel_days_budget_utlised in zip(travel_days_Budget_values, travel_days_budget_utlised):
                        if travel_days_budget.strip() and travel_days_budget_utlised.strip():  # Check if both values are not empty
                            try:
                                # Remove commas before converting to float
                                travel_days_budget_float = float(travel_days_budget.replace(',', ''))
                                travel_days_budget_utlised_float = float(travel_days_budget_utlised.replace(',', ''))
                                travel_days_subtracted_values.append(travel_days_budget_float - travel_days_budget_utlised_float)
                            except ValueError as e:
                                travel_days_subtracted_values.append('')  # or some other value to indicate an error
                        else:
                            travel_days_subtracted_values.append('')  # or some other value to indicate an error
                    # print(f"Subtracted values (Travel Days Budget): {travel_days_subtracted_values}" "185")

                
                # Check the 'DSA Budget' column
                if "DSA Budget" in headers:
                    dsa_index = headers.index("DSA Budget")
                    dsa_budget_values = [str(row[dsa_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                # Check the 'DSA Budget Utlised' column
                if "DSA Budget Utlised" in headers:
                    dsa_index = headers.index("DSA Budget Utlised")
                    dsa_budget_utlised = [str(row[dsa_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values
                    
                    # Subtract DSA Budget Utlised from DSA Budget
                    dsa_subtracted_values = []
                    for dsa_budget, dsa_budget_utlised in zip(dsa_budget_values, dsa_budget_utlised):
                        if dsa_budget.strip() and dsa_budget_utlised.strip():  # Check if both values are not empty
                            try:
                                # Remove commas before converting to float
                                dsa_budget_float = float(dsa_budget.replace(',', ''))
                                dsa_budget_utlised_float = float(dsa_budget_utlised.replace(',', ''))
                                dsa_subtracted_values.append(dsa_budget_float - dsa_budget_utlised_float)
                            except ValueError as e:
                                dsa_subtracted_values.append('')  # or some other value to indicate an error
                        else:
                            dsa_subtracted_values.append('')  # or some other value to indicate an error
                    # print(dsa_subtracted_values)
                    # print("208")

                # Check the 'DSA Days Budget' column
                if "DSA Days Budget" in headers:
                    dsa_days_index = headers.index("DSA Days Budget")
                    dsa_days_Budget_values = [str(row[dsa_days_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                    # Print DSA Days Budget values
                    # print(f"DSA Days Budget values in sheet '{sheet_name}': {dsa_days_Budget_values}","217")

                # Check the 'DSA Days Budget Utlised' column
                if "DSA Days Budget Utlised" in headers:
                    dsa_days_utlised_index = headers.index("DSA Days Budget Utlised")
                    dsa_days_budget_utlised = [str(row[dsa_days_utlised_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                    # Print DSA Days Budget Utlised values
                    # print(f"DSA Days Budget Utlised values in sheet '{sheet_name}': {dsa_days_budget_utlised}","225")

                    # Subtract DSA Days Budget Utlised from DSA Days Budget
                    dsa_days_subtracted_values = []
                    for dsa_days_monthly, dsa_days_budget_utlised in zip(dsa_days_Budget_values, dsa_days_budget_utlised):
                        if dsa_days_monthly.strip() and dsa_days_budget_utlised.strip():  # Check if both values are not empty
                            try:
                                # Remove commas before converting to float
                                dsa_days_monthly_float = float(dsa_days_monthly.replace(',', ''))
                                dsa_days_budget_utlised_float = float(dsa_days_budget_utlised.replace(',', ''))
                                dsa_days_subtracted_values.append(dsa_days_monthly_float - dsa_days_budget_utlised_float)
                            except ValueError as e:
                                # print(f"Error converting values: {e}")
                                dsa_days_subtracted_values.append('')  # or some other value to indicate an error
                        else:
                            dsa_days_subtracted_values.append('')  # or some other value to indicate an error
                    # Print the subtracted values
                    # print(f"Subtracted values (DSA Days Budget): {dsa_days_subtracted_values}","243")

                    # Check the 'Terminal Charges Budget' column
                    if "Terminal Charges Budget" in headers:
                        terminal_charges_index = headers.index("Terminal Charges Budget")
                        terminal_charges_Budget_values = [str(row[terminal_charges_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                        # Print Terminal Charges Budget values
                        # print(f"Terminal Charges Budget values in sheet '{sheet_name}': {terminal_charges_Budget_values}","253")

                    # Check the 'Terminal Charges Budget Utlised' column
                    if "Terminal Charges Budget Utlised" in headers:
                        terminal_charges_utlised_index = headers.index("Terminal Charges Budget Utlised")
                        terminal_charges_budget_utlised = [str(row[terminal_charges_utlised_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                        # Print Terminal Charges Budget Utlised values
                        # print(f"Terminal Charges Budget Utlised values in sheet '{sheet_name}': {terminal_charges_budget_utlised}","261")

                        # Subtract Terminal Charges Budget Utlised from Terminal Charges Budget
                        terminal_charges_subtracted_values = []
                        for terminal_charges_monthly, terminal_charges_budget_utlised in zip(terminal_charges_Budget_values, terminal_charges_budget_utlised):
                            if terminal_charges_monthly.strip() and terminal_charges_budget_utlised.strip():  # Check if both values are not empty
                                try:
                                    # Remove commas before converting to float
                                    terminal_charges_monthly_float = float(terminal_charges_monthly.replace(',', ''))
                                    terminal_charges_budget_utlised_float = float(terminal_charges_budget_utlised.replace(',', ''))
                                    terminal_charges_subtracted_values.append(terminal_charges_monthly_float - terminal_charges_budget_utlised_float)
                                except ValueError as e:
                                    # print(f"Error converting values: {e}")
                                    terminal_charges_subtracted_values.append('')  # or some other value to indicate an error
                            else:
                                terminal_charges_subtracted_values.append('')  # or some other value to indicate an error
                        # Print the subtracted values
                        # print(f"Subtracted values (Terminal Charges Budget): {terminal_charges_subtracted_values}","278")

                        # Check the 'Flight Budget' column
                        if "Flight Budget" in headers:
                            flight_budget_index = headers.index("Flight Budget")
                            flight_budget_values = [str(row[flight_budget_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                            # Print Flight Budget values
                            print(f"Flight Budget values in sheet '{sheet_name}': {flight_budget_values}","286")

                        # Check the 'Flight Budget Utlised' column
                        if "Flight Budget Utlised" in headers:
                            flight_budget_utlised_index = headers.index("Flight Budget Utlised")
                            flight_budget_utlised = [str(row[flight_budget_utlised_index]).replace('\xa0', '').strip() for row in sheet_data]  # Clean values

                            # Print Flight Budget Utlised values
                            print(f"Flight Budget Utlised values in sheet '{sheet_name}': {flight_budget_utlised}","294")

                            # Subtract Flight Budget Utlised from Flight Budget
                            flight_budget_subtracted_values = []
                            for flight_budget_monthly, flight_budget_utlised in zip(flight_budget_values, flight_budget_utlised):
                                if flight_budget_monthly.strip() and flight_budget_utlised.strip():  # Check if both values are not empty
                                    try:
                                        # Remove commas before converting to float
                                        flight_budget_monthly_float = float(flight_budget_monthly.replace(',', ''))
                                        flight_budget_utlised_float = float(flight_budget_utlised.replace(',', ''))
                                        flight_budget_subtracted_values.append(flight_budget_monthly_float - flight_budget_utlised_float)
                                    except ValueError as e:
                                        # print(f"Error converting values: {e}")
                                        flight_budget_subtracted_values.append('')  # or some other value to indicate an error
                                else:
                                    flight_budget_subtracted_values.append('')  # or some other value to indicate an error
                            # Print the subtracted values
                            print(f"Subtracted values (Flight Budget): {flight_budget_subtracted_values}","311")

                # Converting lists to arrays for operations
                pf_Budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in pf_Budget_values]
                pf_subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in pf_subtracted_values]
                travel_Budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in travel_Budget_values]
                travel_subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in travel_subtracted_values]
                travel_days_Budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in travel_days_Budget_values]
                travel_days_subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in travel_days_subtracted_values]
                dsa_budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in dsa_budget_values]
                dsa_subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in dsa_subtracted_values]
                dsa_days_Budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in dsa_days_Budget_values]
                dsa_days_subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in dsa_days_subtracted_values]
                terminal_charges_Budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in terminal_charges_Budget_values]
                terminal_charges_subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in terminal_charges_subtracted_values]
                flight_budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in flight_budget_values]
                flight_budget_subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in flight_budget_subtracted_values]


                pf_Budget_values_array = np.array(pf_Budget_values)
                pf_subtracted_values_array = np.array(pf_subtracted_values)
                travel_Budget_values_array = np.array(travel_Budget_values)
                travel_subtracted_values_array = np.array(travel_subtracted_values)
                travel_days_Budget_values_array = np.array(travel_days_Budget_values)
                travel_days_subtracted_values_array = np.array(travel_days_subtracted_values)
                dsa_budget_values_array = np.array(dsa_budget_values)
                dsa_subtracted_values_array = np.array(dsa_subtracted_values)
                dsa_days_Budget_values_array = np.array(dsa_days_Budget_values)
                dsa_days_subtracted_values_array = np.array(dsa_days_subtracted_values)
                terminal_charges_Budget_values_array = np.array(terminal_charges_Budget_values)
                terminal_charges_subtracted_values_array = np.array(terminal_charges_subtracted_values)
                flight_budget_values_array = np.array(flight_budget_values)
                flight_budget_subtracted_values_array = np.array(flight_budget_subtracted_values)


                #edit on 25th for pf alert and travel budget 

                # Check for the condition for P.F. Budget
                pf_alerts = []
                for i in range(len(pf_Budget_values)):
                    if i < len(consultant_code_values) and pf_Budget_values[i] / 2 >= pf_subtracted_values[i]:
                        alert_message = f"ALERT: Consultant Code {consultant_code_values[i]} P.F. Budget Utlised"
                        pf_alerts.append(alert_message)
                # Print all P.F. Budget alerts collected
                # print(pf_alerts)
             

                # Check for the condition for Travel Budget
                travel_alerts = []
                for i in range(len(travel_Budget_values)):
                    if i < len(consultant_code_values) and travel_Budget_values[i] / 2 >= travel_subtracted_values[i]:
                        alert_message = f"ALERT: Consultant Code {consultant_code_values[i]} Travel Budget Utlised"
                        travel_alerts.append(alert_message)
                # Print all Travel Budget alerts collected
                # print("travel_alerts" "186")
            
                
                # Generate alerts for Travel Days Budget
                travel_days_alerts = []
                for i in range(len(travel_days_Budget_values)):
                    if i < len(consultant_code_values) and travel_days_Budget_values[i] / 2 >= travel_days_subtracted_values[i]:
                        alert_message = f"ALERT: Consultant Code {consultant_code_values[i]} Travel Days Budget Utlised"
                        travel_days_alerts.append(alert_message)
                # print(travel_days_alerts)
                # print("229")

                # Generate alerts for DSA Budget
                dsa_alerts = []
                for i in range(len(dsa_budget_values)):
                    if i < len(consultant_code_values) and dsa_budget_values[i] / 2 >= dsa_subtracted_values[i]:
                        alert_message = f"ALERT: consultant code {consultant_code_values[i]} DSA Budget Utilised"
                        dsa_alerts.append(alert_message)
                # print(dsa_alerts)
                # print(282)

                # Generate alerts for DSA Days Budget
                dsa_days_alerts = []
                for i in range(len(dsa_days_Budget_values)):
                    if i < len(consultant_code_values) and dsa_days_Budget_values[i] / 2 >= dsa_days_subtracted_values[i]:
                        alert_message = f"ALERT: Consultant Code {consultant_code_values[i]} DSA Days Budget Utlised"
                        dsa_days_alerts.append(alert_message)
                # print(dsa_days_alerts)

                # Generate alerts for terminal charge Budget
                terminal_charges_alerts = []
                for i in range(len(terminal_charges_Budget_values)):
                    if i < len(consultant_code_values) and terminal_charges_Budget_values[i] / 2 >= terminal_charges_subtracted_values[i]:
                        alert_message = f"ALERT: Consultant Code {consultant_code_values[i]} terminal charges Budget Utlised"
                        terminal_charges_alerts.append(alert_message)
                # print(terminal_charges_alerts)

                # Generate alerts for flight charge Budget
                flight_alerts = []
                for i in range(len(flight_budget_values)):
                    if i < len(consultant_code_values) and flight_budget_values[i] / 2 >= flight_budget_subtracted_values[i]:
                        alert_message = f"ALERT: Consultant Code {consultant_code_values[i]} flight Budget Utlised"
                        flight_alerts.append(alert_message)
                print(flight_alerts)


                # Replace zero values with numpy.nan
                pf_subtracted_values_array = np.where(pf_subtracted_values_array == 0, np.nan, pf_subtracted_values_array)
                travel_subtracted_values_array = np.where(travel_subtracted_values_array == 0, np.nan, travel_subtracted_values_array)
                
                pf_result = pf_Budget_values_array / pf_subtracted_values_array
                travel_result = travel_Budget_values_array / travel_subtracted_values_array

                pf_result = np.where(np.isnan(pf_result), "Cannot divide by zero", pf_result)
                travel_result = np.where(np.isnan(travel_result), "Cannot divide by zero", travel_result)
                    
            # Prepare context to pass to template
            context = {
                'pf_subtracted_values': pf_subtracted_values,
                'pf_alerts': pf_alerts,  # Combine both alert messages
                'pf_Budget_values': pf_Budget_values,  # Pass P.F. Budget values to template
                'travel_subtracted_values': travel_subtracted_values,
                'travel_alerts': travel_alerts,
                'travel_Budget_values': travel_Budget_values,  # Pass Travel Budget values to template

                'travel_days_alerts': travel_days_alerts,
                'travel_days_subtracted_values': travel_days_subtracted_values,
                'dsa_subtracted_values' : dsa_subtracted_values,
                'dsa_alerts' : dsa_alerts,
                'dsa_days_alerts' : dsa_days_alerts,
                'dsa_days_subtracted_values' : dsa_days_subtracted_values,
                'terminal_charges_alerts' : terminal_charges_alerts,
                'terminal_charges_subtracted_values' : terminal_charges_subtracted_values,
                'flight_alerts' : flight_alerts,
                'flight_budget_subtracted_values' : flight_budget_subtracted_values,

                'excel_data': excel_data,
                
                
             }

            return render(request, 'myapp/index.html', context)

        except Exception as e:
            return HttpResponse(f"Error reading Excel file: {str(e)}")

    return render(request, 'myapp/index.html')
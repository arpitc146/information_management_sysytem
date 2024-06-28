from django.shortcuts import render, redirect
from django.contrib.auth import login,logout,authenticate
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import *
from django.db.models import Q

import pandas as pd
from django.http import HttpResponse
import openpyxl
import numpy as np
# Create your views here.
def loginview(request):
    
    if request.user.is_authenticated:
        return redirect('employees')

    if request.method == 'POST' :
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request,username=username,password=password)
        
        if user is not None:
            login(request,user)
            context ={'user': user}
            return redirect(reverse('employees'),context)
        
        messages.error(request,"Invalid credentials")
        return render(request,"login.html")

    return render(request,"login.html")

@login_required(login_url='login')
def logoutview(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def EmployeeView(request):
    try:
        employees = Employee.objects.all()
        context = {
       
            'employees' : employees
        }
    except:
        pass
    return render(request,'employees.html',context)

@login_required(login_url='login')
def EmployeeProfileView(request,pk):
    flight_budget=0
    travel_budget=0
    ope_budget=0
    employee=0
    remaining_budget_flight = 0
    remaining_budget_ope =0
    remaining_budget_travel = 0
    try:
        employee = Employee.objects.get(uuid=pk)
        flight_budget = FlightBudget.objects.get(employee=employee)
        travel_budget = TravelBudget.objects.get(employee=employee)
        ope_budget = OPEBudget.objects.get(employee=employee)
        remaining_budget_flight = float(flight_budget.allocated_budget) * (0.2)
        remaining_budget_travel = float(travel_budget.allocated_budget) * (0.2)
        remaining_budget_ope = float(ope_budget.allocated_budget) * (0.2)
    except:
        pass

    context = {
        'employee': employee,
        'flight_budget' : flight_budget,
        'travel_budget' : travel_budget,
        'ope_budget' : ope_budget,
        'remaining_budget_flight':remaining_budget_flight,
        'remaining_budget_travel':remaining_budget_travel,
        'remaining_budget_ope':remaining_budget_ope,
    }
    return render(request,"profile.html",context)

@login_required(login_url='login')
def AdvancedTravelPlanView(request,pk):
    try:
        emp = Employee.objects.get(uuid=pk)
        
        atp = AdvancedTravelPlan.objects.filter(employee=emp)
        
    except:
        atp = None

    context = {
        'atp' : atp,
        'emp' : emp
    }
    
    return render(request,"advancedtravelplan.html",context)


@login_required(login_url='login')
def Search(request):
    search_query = request.GET.get('search_query')

    if search_query:
        employ = Employee.objects.filter( Q(Emp_name__icontains=search_query) | Q(contract_no__icontains=search_query))
    else:
        employ = Employee.objects.all()
    context = {
        'employ': employ,
        'search_query': search_query,
        }
    return render(request,'employees.html',context)


@login_required(login_url='login')
def ActualTravelPlan(request,pk):
    try:
        emp = Employee.objects.get(uuid=pk)
        tp = Expense.objects.filter(employee=emp)
        
    except:
        tp = None

    context = {
        'tp' : tp,
        'emp' : emp
    }
    
    return render(request,"actualtravelplan.html",context)

@login_required(login_url='login')
def ViewAtp(request,pk1,pk2):
    try:
        emp = Employee.objects.get(uuid=pk1)
        
        tp = Expense.objects.get(uuid=pk2)
        

    except:
        tp = None
    context = {
        'tp':tp,
        'emp':emp
    }
    return render(request,'viewatp.html',context)

'''

@login_required(login_url='login')
def index(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            return HttpResponse("No file selected for uploading")
        
        file_size = excel_file.size
        print(f"Uploaded file size: {file_size} bytes")

        try:
            # Load the workbook using openpyxl
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active

            # Initialize variables for keyword search
            keyword_locations_pf_monthly = []
            keyword_locations_pf_budget = []

            # Search for 'P.F. Budget' and 'P.F. Budget Utlised' in the worksheet
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and "P.F. Budget".lower() in str(cell.value).lower():
                        keyword_locations_pf_monthly.append((cell.row, cell.column, cell.value))
                        print(f"Found 'P.F. Budget' at row {cell.row}, column {openpyxl.utils.get_column_letter(cell.column)}")
                        print("\n","31")
                    if cell.value and "P.F. Budget Utlised".lower() in str(cell.value).lower():
                        keyword_locations_pf_budget.append((cell.row, cell.column, cell.value))
                        print(f"Found 'P.F. Budget Utlised' at row {cell.row}, column {openpyxl.utils.get_column_letter(cell.column)}")

            # Output the list of found locations for 'P.F. Budget'
            print("Locations found for 'P.F. Budget':")
            for location in keyword_locations_pf_monthly:
                row_num, col_num, cell_value = location
                print(f"Row: {row_num}, Value: {cell_value}")

            # Output the list of found locations for 'P.F. Budget Utlised'
            print("Locations found for 'P.F. Budget Utlised':")
            for location in keyword_locations_pf_budget:
                row_num, col_num, cell_value = location
                print(f"Row: {row_num}, Value: {cell_value}")

            # Read the uploaded Excel file into a Pandas DataFrame
            df = pd.read_excel(excel_file, sheet_name=None, index_col=-1)  # Load all sheets as a dictionary of DataFrames
            sheet1_df = df['Sheet1']
            # Optionally, you can clean up NaN values if needed
            sheet1_df = sheet1_df.dropna(how='all')  # Drop rows where all values are NaN
            # Print the DataFrame for debugging
            print("DataFrame for Sheet1:")
            print(sheet1_df,"57")
            print("\n")
            
            
            # Process each sheet in the workbook
            alert_messages = []
            excel_data = []

            for sheet_name, data in df.items():
                data = data.fillna('')  # Replace NaN with empty strings for display purposes
                # Convert the DataFrame to a list of lists (each list representing a row)
                sheet_data = data.values.tolist()

                # Add headers separately
                headers = data.columns.tolist()
                excel_data.append({"sheet_name": sheet_name, "headers": headers, "data": sheet_data})

                # Check the 'P.F. Budget' column
                if "P.F. Budget" in headers:
                    pf_index = headers.index("P.F. Budget")
                    pf_Budget_values = [str(row[pf_index]).replace('\xa0', '').strip() for row in sheet_data[0:]]  # Skip header row and clean values

                    # Print P.F. Budget values
                    print(f"P.F. Budget values in sheet: {pf_Budget_values}")
                    print(f"P.F. Budget values in sheet '{sheet_name}': {pf_Budget_values}","71")
                   
                   
                # Check the 'Contract No.' column
                if "Contract No." in headers:
                    contract_no_index = headers.index("Contract No.")
                    contract_no_values = [str(row[contract_no_index]).replace('\xa0', '').strip() for row in sheet_data]

                    # Print Contract No. values
                    print(f"Contract No. values in sheet: {contract_no_values}")
                    print(f"Contract No. values in sheet '{sheet_name}': {contract_no_values}","90")
                    
                  # Check the P.F. Budget Utlised  column
                if "P.F. Budget Utlised" in headers:
                    pf_index = headers.index("P.F. Budget Utlised")
                    pf_budget_utlised = [str(row[pf_index]).replace('\xa0', '').strip() for row in sheet_data[0:]]  # Skip header row and clean values

                    # Print P.F. Budget values
                    print(f"P.F. Budget Utlised values in sheet: {pf_budget_utlised}")
                    print(f"P.F. Budget Utlised values in sheet '{sheet_name}': {pf_budget_utlised}","88")
                    # Print the type of pf_budget_utlised list
                    print(f"Type of pf_budget_utlised: {type(pf_budget_utlised)}","90")
                    
                    print("\n")
                    # Subtract P.F. Budget Utlised from P.F. Budget
                    subtracted_values = []
                    for pf_monthly, pf_budget_utlised in zip(pf_Budget_values, pf_budget_utlised):
                        if pf_monthly.strip() and pf_budget_utlised.strip():  # Check if both values are not empty
                            try:
                                # Remove commas before converting to float
                                pf_monthly_float = float(pf_monthly.replace(',', ''))
                                pf_budget_utlised_float = float(pf_budget_utlised.replace(',', ''))
                                subtracted_values.append(pf_monthly_float - pf_budget_utlised_float)
                            except ValueError as e:
                                print(f"Error converting values: {e}")
                                subtracted_values.append('')  # or some other value to indicate an error
                        else:
                            subtracted_values.append('')  # or some other value to indicate an error
                    # Print the subtracted values
                    print(f"Subtracted values: {subtracted_values}")
                    
                    print("\n")
                    
                    # Check the 'Consultant Code' column
                    if "Consultant Code" in headers:
                        consultant_code_index = headers.index("Consultant Code")
                        consultant_code_values = [str(row[consultant_code_index]).replace('\xa0', '').strip() for row in sheet_data]
                        # Print Consultant Code values
                        print(f"Consultant Code values in sheet '{sheet_name}': {consultant_code_values}","130")
                        
                    pf_Budget_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in pf_Budget_values]
                    subtracted_values = [float(x.replace(',', '')) if isinstance(x, str) else x for x in subtracted_values]
                    pf_Budget_values_array = np.array(pf_Budget_values)
                    subtracted_values_array = np.array(subtracted_values)
                    
                    # Check for the condition
                    alerts = []
                    for i in range(len(pf_Budget_values)):
                        if pf_Budget_values[i] / 2 <= subtracted_values[i]:
                            alert_message = f"ALERT: Consultant Code {consultant_code_values[i]} P.F. Budget Utlised"
                            alerts.append(alert_message)
                    # Print all alerts collected
                    print(alerts)
                    print("\n".join(alerts))
                    # Replace zero values with numpy.nan
                    subtracted_values_array = np.where(subtracted_values_array == 0, np.nan, subtracted_values_array)
                    result = pf_Budget_values_array / subtracted_values_array
                    result = np.where(np.isnan(result), "Cannot divide by zero", result)
                    #print(result,"112")
                    


#------------------------------------------------------------------------------------------
                    # Check the 'Travel Budget' column
                    if "Travel Budget" in headers and "Travel Budget Utlised" in headers:
                        travel_budget_index = headers.index("Travel Budget")
                        travel_budget_utlised_index = headers.index("Travel Budget Utlised")

                        travel_budget_values = [str(row[travel_budget_index]).replace('\xa0', '').strip() for row in sheet_data[0:]]
                        travel_budget_utlised = [str(row[travel_budget_utlised_index]).replace('\xa0', '').strip() for row in sheet_data[0:]]

                        print(f"Travel Budget values in sheet: {travel_budget_values}")
                        print(f"Travel Budget Utlised values in sheet: {travel_budget_utlised}")

                        travel_subtracted_values = []
                        for travel_monthly, travel_budget_utlised in zip(travel_budget_values, travel_budget_utlised):
                            if travel_monthly.strip() and travel_budget_utlised.strip():
                                try:
                                    travel_monthly_float = float(travel_monthly.replace(',', ''))
                                    travel_budget_utlised_float = float(travel_budget_utlised.replace(',', ''))
                                    travel_subtracted_values.append(travel_monthly_float - travel_budget_utlised_float)
                                except ValueError as e:
                                    print(f"Error converting values: {e}")
                                    travel_subtracted_values.append('')
                            else:
                                travel_subtracted_values.append('')

                        print(f"Travel subtracted values: {travel_subtracted_values}")
                        
                        #For alert msg 
                        travel_alerts = []
                        travel_budget_values = [float(str(x).replace(',', '')) for x in travel_budget_values]
                        travel_subtracted_values = [float(str(x).replace(',', '')) for x in travel_subtracted_values]
                        consultant_code_values = [str(x) for x in consultant_code_values]
                        for i in range(len(travel_budget_values)):
                            if travel_budget_values[i] / 2 <= travel_subtracted_values[i]:
                                alert_message_travel = f"ALERT: Consultant Code {consultant_code_values[i]} Travel Budget Utilised"
                                print("\n")
                                print(alert_message_travel,"187")
                                travel_alerts.append(consultant_code_values[i])

                        print(travel_alerts, "189")
                        
#------------------------------------------------------------------------------------------

                       

                    
#------------------------------------------------------------------------------------------
            # Prepare context to pass to template
            context = {
                'subtracted_values': subtracted_values,
                'excel_data': excel_data,
                'alert_messages': alerts,
                'pf_Budget_values': pf_Budget_values,  # Pass P.F. Budget values to template
                'travel_alerts': travel_alerts,
                'travel_budget_values': travel_budget_values,  # Pass Travel Budget values to template
            }


            return render(request, 'index.html', context)

        except Exception as e:
            return HttpResponse(f"Error reading Excel file: {str(e)}")

    return render(request, 'index.html')
    '''
        

